import pandas as pd
from datetime import time
import streamlit as st
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

st.set_page_config(page_title="Captain wise Sales", layout="wide")

# Custom CSS for colors
st.markdown("""
<style>
    /* Grand total row - light green background */
    .grandtotal-row {
        background-color: #d4edda !important;
        font-weight: bold !important;
    }
    /* Item rows - distinct colors per captain */
    .captain-0 {
        background-color: #e6ffe6 !important;  /* Light green */
    }
    .captain-1 {
        background-color: #e6f0ff !important;  /* Light blue */
    }
    .captain-2 {
        background-color: #fff2e6 !important;  /* Light orange */
    }
    .captain-3 {
        background-color: #ffe6f0 !important;  /* Light pink */
    }
    .captain-4 {
        background-color: #f0e6ff !important;  /* Light purple */
    }
    /* Summary section styling */
    .item-card {
        background-color: white;
        padding: 10px;
        border-radius: 5px;
        box-shadow: 0 1px 3px rgba(0,0,0,0.1);
        margin: 5px;
        text-align: center;
    }
</style>
""", unsafe_allow_html=True)

st.title("📊 Captain wise Sales")
st.write("Upload your CSV file to generate the report")

# File upload
uploaded_file = st.file_uploader("Choose a CSV file", type="csv")

if uploaded_file is not None:
    try:
        # Read the uploaded file
        df = pd.read_csv(uploaded_file, skiprows=5)
        st.success("File uploaded successfully!")
        
        with st.spinner("Processing..."):
            # Filter data
            file1 = df.iloc[:-1] if len(df) > 0 else df
            file2 = file1[file1['Order Type'] == 'Dine-In'].copy()
            file = file2[['Item Name', 'Order Type', 'Quantity', 'Captain Name','Billed Time']]
            file = file[file['Captain Name'].str.lower() != 'without_captain'].copy()
            
            # Convert Billed Time
            file['Billed Time'] = pd.to_datetime(file['Billed Time'])
            
            # Time categorization
            def get_time_category(billed_time):
                t = billed_time.time()
                if time(6, 0) <= t <= time(11, 00):
                    return 'Breakfast'
                elif time(11, 1) <= t <= time(16, 0):
                    return 'Lunch'
                elif time(16, 1) <= t <= time(18, 0):
                    return 'Snacks'
                else:
                    return 'Late Night'
            
            file['Time Category'] = file['Billed Time'].apply(get_time_category)
            
            # Group items
            def Grouped_items(ItemName):
                item_name = str(ItemName).lower()
                if "tamil nadu meals" in item_name:
                    return "Tamilnadu Meals"
                elif 'curd rice' in item_name:
                    return "Classic Curd Rice"
                elif 'thali' in item_name:
                    return "North Indian Thali"
                elif "special soup" in item_name:
                    return "Day Spl Soup"
                elif "tandoori platter" in item_name:
                    return "Tandoori Platter"
                elif "kunafa" in item_name:
                    return 'Kunafa_item'
                elif "fried rice" in item_name:
                    return 'Fried rice'
                elif "noodles" in item_name:
                    return 'Noodles_Item'
                elif "falooda" in item_name:
                    return "Falooda_Item"
                else:
                    return None
            
            file['Item Name'] = file['Item Name'].apply(Grouped_items)
            file = file.dropna(subset=['Item Name']).copy()
            
            # Create report
            unique_captains = sorted(file['Captain Name'].unique())
            
            # Create a mapping of captain to color index
            captain_to_color = {}
            for idx, captain in enumerate(unique_captains):
                captain_to_color[captain] = idx % 5  # Cycle through 5 colors
            
            # Track grand totals
            grand_total_breakfast = 0
            grand_total_lunch = 0
            grand_total_snacks = 0
            grand_total_latenight = 0
            grand_total_all = 0
            
            # Track item-wise totals
            item_totals = {}
            
            # Create a list to store all rows for display
            display_rows = []
            
            # Process each captain
            for captain in unique_captains:
                captain_data = file[file['Captain Name'] == captain]
                captain_items = sorted(captain_data['Item Name'].unique())
                color_index = captain_to_color[captain]
                
                # Add captain name row (no color)
                display_rows.append({
                    'Captain Name': captain,
                    'Item Name': '',
                    'Breakfast': '',
                    'Lunch': '',
                    'Snacks': '',
                    'Late Night': '',
                    'Total': ''
                })
                
                # Process each item for this captain
                for item in captain_items:
                    item_data = captain_data[captain_data['Item Name'] == item]
                    pivot = pd.pivot_table(
                        item_data,
                        values='Quantity',
                        columns='Time Category',
                        aggfunc='sum',
                        fill_value=0
                    )
                    
                    # Get values
                    breakfast = int(pivot.get('Breakfast', 0))
                    lunch = int(pivot.get('Lunch', 0))
                    snacks = int(pivot.get('Snacks', 0))
                    late_night = int(pivot.get('Late Night', 0))
                    total = breakfast + lunch + snacks + late_night
                    
                    # Add to grand totals
                    grand_total_breakfast += breakfast
                    grand_total_lunch += lunch
                    grand_total_snacks += snacks
                    grand_total_latenight += late_night
                    grand_total_all += total
                    
                    # Add to item-wise totals
                    if item not in item_totals:
                        item_totals[item] = 0
                    item_totals[item] += total
                    
                    # Add item row with color based on captain
                    display_rows.append({
                        'Captain Name': '',
                        'Item Name': item,
                        'Breakfast': breakfast if breakfast > 0 else '',
                        'Lunch': lunch if lunch > 0 else '',
                        'Snacks': snacks if snacks > 0 else '',
                        'Late Night': late_night if late_night > 0 else '',
                        'Total': total,
                        'color_class': f'captain-{color_index}'  # Add color class for styling
                    })
            
            # Add grand total row
            display_rows.append({
                'Captain Name': 'GRAND TOTAL',
                'Item Name': '--- ALL ITEMS TOTAL ---',
                'Breakfast': grand_total_breakfast if grand_total_breakfast > 0 else '',
                'Lunch': grand_total_lunch if grand_total_lunch > 0 else '',
                'Snacks': grand_total_snacks if grand_total_snacks > 0 else '',
                'Late Night': grand_total_latenight if grand_total_latenight > 0 else '',
                'Total': grand_total_all,
                'color_class': 'grandtotal-row'
            })
            
            # Create dataframe
            df_display = pd.DataFrame(display_rows)
            
            # Display results
            st.subheader("📋 Processed Results")
            
            # Apply CSS classes for coloring
            def color_rows(row):
                if 'color_class' in row and row['color_class'] == 'grandtotal-row':
                    return ['background-color: #d4edda; font-weight: bold'] * len(row)
                elif 'color_class' in row and row['color_class']:
                    # This is an item row with captain color
                    color_map = {
                        'captain-0': 'background-color: #e6ffe6',
                        'captain-1': 'background-color: #e6f0ff',
                        'captain-2': 'background-color: #fff2e6',
                        'captain-3': 'background-color: #ffe6f0',
                        'captain-4': 'background-color: #f0e6ff',
                    }
                    color = color_map.get(row['color_class'], '')
                    return [color] * len(row)
                else:
                    # Captain name row - no color
                    return [''] * len(row)
            
            # Apply styling to the dataframe
            if 'color_class' in df_display.columns:
                styled_df = df_display.drop('color_class', axis=1).style.apply(color_rows, axis=1)
            else:
                styled_df = df_display.style
            
            # Display the dataframe
            st.dataframe(styled_df, use_container_width=True, height=500)
            
            # Item-wise summary
            st.subheader("📦 Item-wise Total Quantity")
            
            # Display items in multiple columns
            cols = st.columns(3)
            items_list = sorted(item_totals.items(), key=lambda x: x[1], reverse=True)
            for idx, (item, qty) in enumerate(items_list):
                col_idx = idx % 3
                with cols[col_idx]:
                    st.markdown(f"""
                    <div class="item-card">
                        <strong>{item}</strong><br>
                        <span style="font-size: 24px; color: #FF4B4B;">{qty}</span>
                    </div>
                    """, unsafe_allow_html=True)
            
            # Time-wise summary
            st.subheader("⏰ Time-wise Summary")
            col1, col2, col3, col4, col5 = st.columns(5)
            
            with col1:
                st.metric("Total Items", grand_total_all)
            with col2:
                st.metric("Breakfast", grand_total_breakfast)
            with col3:
                st.metric("Lunch", grand_total_lunch)
            with col4:
                st.metric("Snacks", grand_total_snacks)
            with col5:
                st.metric("Late Night", grand_total_latenight)
            
            # Download buttons
            st.subheader("💾 Download")
            col1, col2 = st.columns(2)
            
            with col1:
                # Create Excel file
                output = BytesIO()
                
                # Create a workbook
                wb = Workbook()
                ws = wb.active
                ws.title = "Sales Data"
                
                # Define colors for Excel
                colors = {
                    'grand_total': PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid'),
                    0: PatternFill(start_color='E6FFE6', end_color='E6FFE6', fill_type='solid'),
                    1: PatternFill(start_color='E6F0FF', end_color='E6F0FF', fill_type='solid'),
                    2: PatternFill(start_color='FFF2E6', end_color='FFF2E6', fill_type='solid'),
                    3: PatternFill(start_color='FFE6F0', end_color='FFE6F0', fill_type='solid'),
                    4: PatternFill(start_color='F0E6FF', end_color='F0E6FF', fill_type='solid'),
                }
                
                # Add headers
                headers = ['Captain Name', 'Item Name', 'Breakfast', 'Lunch', 'Snacks', 'Late Night', 'Total']
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = Border(
                        left=Side(style='thin'), 
                        right=Side(style='thin'),
                        top=Side(style='thin'), 
                        bottom=Side(style='thin')
                    )
                
                # Add data with colors
                for row_num, row_data in enumerate(display_rows, 2):
                    # Determine row color (only for items and grand total)
                    if 'color_class' in row_data:
                        if row_data['color_class'] == 'grandtotal-row':
                            row_color = colors['grand_total']
                            font = Font(bold=True)
                        elif row_data['color_class'].startswith('captain-'):
                            color_index = int(row_data['color_class'].split('-')[1])
                            row_color = colors.get(color_index, colors[0])
                            font = Font(bold=False)
                        else:
                            row_color = None
                            font = Font(bold=False)
                    else:
                        # Captain name row - no color
                        row_color = None
                        font = Font(bold=False)
                    
                    for col_num, col_name in enumerate(headers, 1):
                        cell = ws.cell(row=row_num, column=col_num)
                        cell.value = row_data[col_name]
                        if row_color:
                            cell.fill = row_color
                        cell.font = font
                        cell.alignment = Alignment(horizontal='center' if col_name in ['Breakfast', 'Lunch', 'Snacks', 'Late Night', 'Total'] else 'left')
                        cell.border = Border(
                            left=Side(style='thin'), 
                            right=Side(style='thin'),
                            top=Side(style='thin'), 
                            bottom=Side(style='thin')
                        )
                
                # Auto-fit columns
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 30)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                wb.save(output)
                
                st.download_button(
                    label="📥 Download Excel (Formatted)",
                    data=output.getvalue(),
                    file_name="processed_sales_data.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            
            with col2:
                # Prepare CSV data (remove color_class)
                csv_df = pd.DataFrame([
                    {k: v for k, v in row.items() if k != 'color_class'}
                    for row in display_rows
                ])
                csv_data = csv_df.to_csv(index=False).encode('utf-8')
                st.download_button(
                    label="📥 Download CSV",
                    data=csv_data,
                    file_name="processed_sales_data.csv",
                    mime="text/csv",
                    use_container_width=True
                )
                    
    except Exception as e:
        st.error(f"Error: {str(e)}")
